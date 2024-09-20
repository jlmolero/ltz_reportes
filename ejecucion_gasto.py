import pandas as pd
import datetime as dt
import openpyxl
import re
#from openpyxl.styles import Font, Alignment


meses_interes = [7] # mes que me interesa, en este caso julio

meses={
    'enero': 1,
    'febrero': 2,
    'marzo': 3,
    'abril': 4,
    'mayo': 5,
    'junio': 6,
    'julio': 7,
    'agosto': 8,
    'septiembre': 9,
    'octubre': 10,
    'noviembre': 11,
    'diciembre': 12
}

if meses_interes == [1, 2, 3]:
    periodo = 'Primer Trimestre'
elif meses_interes == [4, 5, 6]:
    periodo = 'Segundo Trimestre'
elif meses_interes == [7, 8, 9]:
    periodo = 'Tercer Trimestre'
elif meses_interes == [10, 11, 12]:
    periodo = 'Cuarto Trimestre'
else:
    periodo=''
    for mes, num in meses.items():
        if num in meses_interes:
            periodo=periodo + mes + ', '
#    periodo = [mes for mes, num in meses.items() if num in meses_interes]


#Ruta de Archivo origen de pagos
ruta_archivo = "./data/RegistroPagos.xlsx"
#ruta_archivo = r"\\dc2\administracion\2024\CUADROS\Registro de Pagos.xlsx"



#cargar los datos en el dataframe
datos = pd.read_excel(ruta_archivo, sheet_name="Pagos")


#Eliminar Ordenes de Pago Anuladas
datos = datos[datos['NroFactura']!='ANULADA']

#Eliminar registro donde 'OrdenPago' es nulo o vacio
datos = datos[datos['OrdenPago'].notnull()]

#Eliminar filas de impuestos, pues estas no reflejan la ejecucion de gastos, sino más bien el pago de lo retenido
datos = datos[datos['CodigoPartida']!='4.03.18.01.00']
datos = datos[datos['CodigoPartida']!='4.03.18.99.00']

#Eliminar columnas innecesarias
# el atributo inplace indica si se va a modificar el dataframe original o se va a crear un nuevo dataframe
# con los valores modificados. Si inplace es True, se modifica el original, si es False, se crea un nuevo dataframe
# el atributo axis indica si se va a eliminar la columna (axis=0) o la fila (axis=1) que coincide con el valor especificado

datos.drop('NotaDebito', inplace=True, axis=1) # se elimina la columna 'NotaDebito' en el dataframe original
datos.drop('Referencia', inplace=True, axis=1) # se elimina la columna 'Referencia' en el dataframe original
datos.drop('NroFactura', inplace=True, axis=1)
datos.drop('MesDeclaracion', inplace=True, axis=1)
datos.drop('PeriodoDeclaracion', inplace=True, axis=1)
datos.drop('ProcesoNro', inplace=True, axis=1)



#Rellenar los registros de FuenteDeFinanciamiento vacios según la Cuenta
# Define un diccionario que mapea los valores de la columna Cuenta a los valores de la columna FuenteDeFinanciamiento
cuenta_fuente_financiamiento = {
    '4363': 'Recursos por Operaciones',
    '6597': 'Recursos por Operaciones',
    '0171': 'Situado Constitucional',
    'PATRIA': 'Situado Constitucional'
}

partida_descripcion = {
    '4.01':'GASTOS DE PERSONAL',
    '4.02':'MATERIALES, SUMINISTROS Y MERCANCÍAS',
    '4.03':'SERVICIOS NO PERSONALES',
    '4.04':'ACTIVOS REALES',
    '4.05':'ACTIVOS  FINANCIEROS',
    '4.06':'GASTOS DE DEFENSA Y SEGURIDAD DEL ESTADO',
    '4.07':'TRANSFERENCIAS Y DONACIONES',
    '4.08':'OTROS GASTOS',
    '4.09':'ASIGNACIONES NO DISTRIBUIDAS',
    '4.10':'SERVICIO DE LA DEUDA PÚBLICA',
    '4.11':'DISMINUCIÓN DE PASIVOS',
    '4.12':'DISMINUCIÓN DEL PATRIMONIO',
    '4.98':'RECTIFICACIONES AL PRESUPUESTO'
}

# Utiliza el método map para aplicar el diccionario a la columna Cuenta y obtener un nuevo objeto Series con los resultados
fuente_financiamiento = datos['Cuenta'].map(cuenta_fuente_financiamiento)

# Rellena los valores vacíos en la columna FuenteDeFinanciamiento con los valores obtenidos en el paso anterior
#datos['FuenteDeFinanciamiento'].fillna(fuente_financiamiento, inplace=True)
datos = datos.assign(FuenteDeFinanciamiento=datos['FuenteDeFinanciamiento'].fillna(fuente_financiamiento))

# Crea un nuevo dataframe que incluya solo las columnas necesarias para el iva
datosIVA = datos[['N', 'Fecha', 'Cuenta', 'OrdenPago', 'DocBeneficiario', 'Beneficiario', 'Descripcion', 'IVA', 'FuenteDeFinanciamiento']].copy()
datosIVA['IVA'] = datosIVA['IVA'].fillna(0)
datosIVA['IVA'] = datosIVA['IVA'].apply(lambda x: 0 if x in ['-',''] else x)




# Ahora quiero crear un nuevo dataframe llamado 'informe_gasto_ejecutado' que agrupe los registros según las columnas
# 'CodigoPartida' y 'DescripcionPartida'
# Me interesa que se muestren los valores de 'MontoSinIVA' tanto totales, como disgregados según el valor de
# 'FuenteDeFinanciamiento'
# Para lograr esto, voy a utilizar el método groupby, que me permite agrupar los registros de acuerdo a las
# columnas especificadas en el parámetro by

datos['Mes'] = datos['Fecha'].dt.month  # extrae el mes de cada fecha
datos = datos[datos['Mes'].isin(meses_interes)]  # filtra los datos según la lista de meses

informe_gasto_ejecutado = datos.groupby(['CodigoPartida', 'DescripcionPartida', 'FuenteDeFinanciamiento'])['MontoSinIVA'].sum().unstack('FuenteDeFinanciamiento').fillna(0).rename_axis(None, axis=1).reset_index()

#informe_gasto_ejecutado = datos[datos['Fecha'].dt.month == mes_interes].groupby(['CodigoPartida', 'DescripcionPartida', 'FuenteDeFinanciamiento'])['MontoSinIVA'].sum().unstack('FuenteDeFinanciamiento').fillna(0).rename_axis(None, axis=1).reset_index()


#Calcular el IVA del mes
#iva_mes = datosIVA[datosIVA['Fecha'].dt.month == mes_interes].pivot_table(values='IVA', index=None, columns='FuenteDeFinanciamiento', aggfunc='sum').fillna(0).reset_index()
datosIVA['Mes'] = datosIVA['Fecha'].dt.month  # extrae el mes de cada fecha
datosIVA = datosIVA[datosIVA['Mes'].isin(meses_interes)]  # filtra los datos según la lista de meses

iva_mes = datosIVA[datosIVA['Mes'].isin(meses_interes)].pivot_table(values='IVA', index=None, columns='FuenteDeFinanciamiento', aggfunc='sum').fillna(0).reset_index()

#Crear Fila para agregar en el informe
fila_iva_mes = pd.DataFrame({'CodigoPartida':['4.03.18.01.00'], 'DescripcionPartida':['Impuesto al valor Agregado'], 'Recursos por Operaciones':[iva_mes.loc[0, 'Recursos por Operaciones']], 'Situado Constitucional':[iva_mes.loc[0, 'Situado Constitucional']]}).fillna(0)

informe_gasto_ejecutado = pd.concat([informe_gasto_ejecutado, fila_iva_mes], ignore_index=True).sort_values(by='CodigoPartida')

# Ahora voy a agregar una columna con el valor total de 'MontoSinIVA' para cada agrupación
informe_gasto_ejecutado['Total'] = informe_gasto_ejecutado.iloc[:, 2:].sum(axis=1)

#Calcular subtotales
informe_gasto_ejecutado_subtotales = informe_gasto_ejecutado.groupby(informe_gasto_ejecutado['CodigoPartida'].str[:4]).agg({'Recursos por Operaciones':'sum', 'Situado Constitucional':'sum', 'Total': 'sum'}).rename_axis('CodigoPartida').reset_index()
informe_gasto_ejecutado_subtotales['CodigoPartida'] = informe_gasto_ejecutado_subtotales['CodigoPartida']# + '.00.00.00'
informe_gasto_ejecutado_subtotales['DescripcionPartida'] = informe_gasto_ejecutado_subtotales['CodigoPartida'].map(partida_descripcion)


#fila de totales generales
fila_totales_generales = pd.DataFrame({'CodigoPartida':['TOTALES'], 'DescripcionPartida':[''], 
                                      'Recursos por Operaciones':[informe_gasto_ejecutado['Recursos por Operaciones'].sum()], 
                                      'Situado Constitucional':[informe_gasto_ejecutado['Situado Constitucional'].sum()],
                                      'Total':[informe_gasto_ejecutado['Total'].sum()]})


# Agregar filas de subototales y totales generales
informe_gasto_ejecutado = pd.concat([informe_gasto_ejecutado, informe_gasto_ejecutado_subtotales], ignore_index=True).sort_values(by='CodigoPartida')
informe_gasto_ejecutado = pd.concat([informe_gasto_ejecutado, fila_totales_generales], ignore_index=True)

print(informe_gasto_ejecutado)
print(periodo)


"""
##########################
######### Exportar a Excel
##########################
if len(meses_interes) == 1:
    informe_gasto_ejecutado.to_excel('informe_gasto_ejecutado_{}.xlsx'.format(meses_interes[0]), index=False)
elif len(meses_interes) > 1:
    period=''
    for mes in meses_interes:
        period=period + '_' + str(mes)
    informe_gasto_ejecutado.to_excel(f'informe_gasto_ejecutado_{period}.xlsx', index=False)


##########################
####### Dar Formato al Iinforme
##########################

import openpyxl

if len(meses_interes) == 1:
    archivo_excel=f'informe_gasto_ejecutado_{meses_interes[0]}.xlsx'
elif len(meses_interes) > 1:
    period=''
    for mes in meses_interes:
        period=period + '_' + str(mes)
    archivo_excel = f'informe_gasto_ejecutado_{period}.xlsx'
#else:
#    archivo_excel = 'informe_gasto_ejecutado.xlsx'

wb = openpyxl.load_workbook(archivo_excel)
sheet = wb.active

#Definir propiedades de la hoja
sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
sheet.page_setup.fitToHeight = 0
sheet.page_setup.fitToWidth = 0
sheet.oddFooter.center.text = 'Pág. &P de &N'
#Margenes
conversor=0.3937007874 # 1cm = 0.3937007874 in
sheet.page_margins.top = (1.9 * conversor)
sheet.page_margins.bottom = (1.9 * conversor)
sheet.page_margins.left = (0.6 * conversor)
sheet.page_margins.right = (0.6 * conversor)
sheet.page_margins.header = (0.8 * conversor)
sheet.page_margins.footer = (0.8 * conversor)
#Repetir primeras 5 filas
sheet.print_title_rows = '1:5'

#Definir Ancho de Columnas
sheet.column_dimensions['A'].width = 13.71
sheet.column_dimensions['B'].width = 44.14
sheet.column_dimensions['C'].width = 13.71
sheet.column_dimensions['D'].width = 13.71
sheet.column_dimensions['E'].width = 13.71

#Insertar filas vacias al principios del documento
sheet.insert_rows(1, 4)


#Definir altura de las filas
for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False), start=1):
    if any(cell.value is not None for cell in row):
        sheet.row_dimensions[idx].height = 30

#Escribir Titulo
sheet.merge_cells('A1:E1')
sheet['A1'].alignment = Alignment(horizontal='center')
sheet['A1'] = 'INFORME DE GASTO EJECUTADO'
sheet['A1'].font = Font(name='Calibri', size=16, bold=True)

#Escribir Periodo
sheet.row_dimensions[2].height = 5
sheet.merge_cells('D3:E3')
sheet['C3'].alignment = Alignment(horizontal='right')
sheet['C3'].font = Font(name='Calibri', size=12, bold=False)
sheet['C3'] = 'PERIODO: '
sheet['D3'].font = Font(name='Calibri', size=12, bold=True)
sheet['D3'] = f"{periodo} 2024"

#Definir que la fila de titulos sea negrita y que el texto se ajuste 
for row in sheet['A5:E5']:
    for cell in row:
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, values_only=False):
    #Filas de Subtotales de Partidas
    if isinstance(row[0].value, str) and re.match(r'^\d\.\d{2}$', row[0].value): #len(row[0].value) == 4:
        row[1].alignment = Alignment(wrap_text=True)
        row[2].number_format = '#,##0.00'
        row[3].number_format = '#,##0.00'
        row[4].number_format = '#,##0.00'
        
        for cell in row:
            cell.border = openpyxl.styles.borders.Border(
                top=openpyxl.styles.borders.Side(style='thin'),
                bottom=openpyxl.styles.borders.Side(style='thin'),
                left=openpyxl.styles.borders.Side(style='thin'),
                right=openpyxl.styles.borders.Side(style='thin'))
            cell.fill = openpyxl.styles.PatternFill(fgColor='D9D9D9', fill_type='solid')
            cell.font = Font(name='Calibri', size=12, bold=True)
    #filas Normales
    elif isinstance(row[0].value, str) and re.match(r'^\d\.\d{2}\.\d{2}\.\d{2}\.\d{2}$', row[0].value):
        row[1].alignment = Alignment(wrap_text=True)
        row[2].number_format = '#,##0.00'
        row[3].number_format = '#,##0.00'
        row[4].number_format = '#,##0.00'
        
        for cell in row:
            cell.border = openpyxl.styles.borders.Border(
                top=openpyxl.styles.borders.Side(style='thin'),
                bottom=openpyxl.styles.borders.Side(style='thin'),
                left=openpyxl.styles.borders.Side(style='thin'),
                right=openpyxl.styles.borders.Side(style='thin'))
            cell.font = Font(name='Calibri', size=11)

    #Fila Totales
    elif isinstance(row[0].value, str) and row[0].value=='TOTALES':
        row[2].number_format = '#,##0.00'
        row[3].number_format = '#,##0.00'
        row[4].number_format = '#,##0.00'
        
        for cell in row:
            cell.border = openpyxl.styles.borders.Border(
                top=openpyxl.styles.borders.Side(style='thin'),
                bottom=openpyxl.styles.borders.Side(style='thin'),
                left=openpyxl.styles.borders.Side(style='thin'),
                right=openpyxl.styles.borders.Side(style='thin'))
            cell.fill = openpyxl.styles.PatternFill(fgColor='404040', fill_type='solid')
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')



wb.save(archivo_excel)
"""