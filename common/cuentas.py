import os
from os import path
import openpyxl
import re
import pandas as pd
import yaml
from common.basicas import latino_a_numero
import platform

######################
#
# Cuentas
#
######################
if platform.system() == 'Windows':
    with open('./config.yaml', 'r') as f:
        config = yaml.safe_load(f)
elif platform.system() == 'Linux':
    with open('./config_linux.yaml', 'r') as f:
        config = yaml.safe_load(f)

cuentas = {}
for cuenta, valores in config['cuentas'].items():
    cuentas[cuenta] = {
        'ruta': valores['ruta'],
        'banco': valores['banco']
    }

######################
# FUNCION PARA OBTENER LOS MOVIMIENTOS DE UNA CUENTA
######################
def get_movimientos(cuenta, meses_interes=None):
    ruta = cuentas[cuenta]['ruta']
    banco = cuentas[cuenta]['banco']
    movimientos = pd.DataFrame()
    for file in os.listdir(ruta):
        if file.endswith(".xlsx"):
            wb = openpyxl.load_workbook(os.path.join(ruta, file))
            for sheet in wb:
                if sheet.title != 'Resumen':
                    df = pd.read_excel(os.path.join(ruta, file), sheet_name=sheet.title)
                    movimientos = pd.concat([movimientos, df], ignore_index=True)
    #
    # PREPARAR MOVIEMIENTOS
    #
    if banco == 'BANCRECER':
        # convertir la columna 'FECHA' a tipo datetime
        movimientos['FECHA'] = pd.to_datetime(movimientos['FECHA'], dayfirst=True)

        #Seleccionar solo los movimientos correspondiente al periodo deseado
        if meses_interes is not None:
            movimientos = movimientos[movimientos['FECHA'].dt.month.isin(meses_interes)]

        # preparar la columna REFERENCIA
        movimientos['REFERENCIA'] = movimientos['REFERENCIA'].astype(str).str.zfill(8)

        #Crear La Columna LOTE cuyos valores serán los mismos de la columna REFERENCIA
        movimientos['lote'] = movimientos['REFERENCIA']

        #Crear la columna tipoOperacion
        movimientos['tipoOperacion'] = ''
        for index, row in movimientos.iterrows():
            if row['DESCRIPCION']=="N/D CREDITO INM.OB-G200076496":
                movimientos.at[index, 'tipoOperacion'] = 'TRASFERENCIA LTZ OTROS BANCOS'
            elif row['DESCRIPCION'].startswith('N/D CREDITO INM.'):
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS A PROVEEDORES'
            elif row['DESCRIPCION'].startswith('COM.CREDITO INM.OB'):
                movimientos.at[index, 'tipoOperacion'] = 'COMISIONES PAGOS A PROVEEDORES'
            elif row['DESCRIPCION'].startswith('N/C CREDITO INM.'):
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS RECIBIDOS'
            elif row['DESCRIPCION'].startswith('COMISION'):
                movimientos.at[index, 'tipoOperacion'] = 'COMISIONES VARIAS'
            elif row['DESCRIPCION']=="COM. POR EMISION EDO. DE CTA.":
                movimientos.at[index, 'tipoOperacion'] = 'COMISIONES VARIAS'
            elif row['DESCRIPCION']=="COM. POR MANTENIMIENTO DE CTA.":
                movimientos.at[index, 'tipoOperacion'] = 'COMISIONES VARIAS'


        #Convertir las columnas 'DEBITOS' y 'CREDITOS' a tipo float
        movimientos['DEBITOS'] = movimientos['DEBITOS'].apply(latino_a_numero)
        movimientos['CREDITOS'] = movimientos['CREDITOS'].apply(latino_a_numero)
        movimientos['SALDO'] = movimientos['SALDO'].apply(latino_a_numero)
        
        #Crear 2 nuevas columnas, una llamada monto y otra llama tipoMovimiento
        movimientos['monto'] = 0.0
        movimientos['tipoMovimiento'] = ''
        for index, row in movimientos.iterrows():
            if row['DEBITOS'] > 0 and row['CREDITOS'] == 0:
                movimientos.loc[index, 'monto'] = -row['DEBITOS']
                movimientos.loc[index, 'tipoMovimiento'] = 'Nota de Débito'
            elif row['DEBITOS'] == 0 and row['CREDITOS'] > 0:
                movimientos.loc[index, 'monto'] = row['CREDITOS']
                movimientos.loc[index, 'tipoMovimiento'] = 'Nota de Crédito'
        
        #Eliminar las columnas 'DEBITOS' y 'CREDITOS'
        movimientos = movimientos.drop(['DEBITOS', 'CREDITOS'], axis=1)
        
        #Cambiar el nombre de las columnas
        movimientos = movimientos.rename(columns={'FECHA': 'fecha', 'SALDO': 'saldo', 'REFERENCIA': 'referencia', 'DESCRIPCION': 'concepto'})

        #reordenar las columnas
        movimientos = movimientos[['fecha', 'referencia', 'lote', 'concepto', 'monto', 'saldo', 'tipoOperacion']]

    elif banco == 'BDV':
        # elmiar las columnas 'rif' y 'numeroCuenta'
        movimientos = movimientos.drop(['rif', 'numeroCuenta'], axis=1)

        # Eliminar todos los registros cuyo valor para la columnas 'concepto' sea "SALDO INICIAL"
        movimientos = movimientos[movimientos['tipoMovimiento'] != 'Saldo Inicial']

        # convertir la columna 'fecha' a tipo datetime
        movimientos['fecha'] = pd.to_datetime(movimientos['fecha'], dayfirst=True)

        #Seleccionar solo los movimientos correspondiente al periodo deseado
        if meses_interes is not None:
            movimientos = movimientos[movimientos['fecha'].dt.month.isin(meses_interes)]
        
        #Convertir las columnas de 'monto' y saldo' a tipo float
        movimientos['monto'] = movimientos['monto'].apply(latino_a_numero)
        movimientos['saldo'] = movimientos['saldo'].apply(latino_a_numero)

        #Extraer los ultimos 8 digitos de la columna 'referencia' y convertirlos a tipo str
        movimientos['referencia'] = movimientos['referencia'].astype(str).apply(lambda x: x[:-1] if x.endswith(' ') else x)
        movimientos['referencia'] = movimientos['referencia'].str[-8:].str.zfill(8)

        #Crear la columna 'lote'
        movimientos['lote'] = ''
        movimientos['lote'] = movimientos.apply(lambda row: row['concepto'][-8:] if 'LOTE' in row['concepto'] else row['referencia'], axis=1)

        #crear la columna tipoOperacion
        movimientos['tipoOperacion'] = ''
        for index, row in movimientos.iterrows():
            if row['concepto']=="PAGO RECIBIDO OTROS BANCOS 0168 G200076496": #Transferencias internas desde Bancrecer
                movimientos.at[index, 'tipoOperacion'] = 'TRANSFERENCIA RECIBIDA LTZ BANCRECER'
            elif row['concepto'].startswith("PAGO RECIBIDO BDV G200076496"): #Transferencias internas desde BDV
                movimientos.at[index, 'tipoOperacion'] = 'TRANSFERENCIA RECIBIDA LTZ BDV'
            elif row['concepto'].startswith('COMISION PAGO A PROVEEDORES'): # Comisiones de Pagos a Proveedores
                movimientos.at[index, 'tipoOperacion'] = 'COMISIONES PAGOS A PROVEEDORES'
            elif row['concepto'].startswith('PAGO A PROVEEDORES'): # Pagos a Proveedores
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS A PROVEEDORES'
            elif row['concepto'].startswith('COM MANTENIMIENTO DE CUENTA'): # Comisiones de Mantenimiento de cuenta
                movimientos.at[index, 'tipoOperacion'] = 'COMISIONES VARIAS'
            elif row['concepto'].startswith("PAGO IMPUESTOS INTERNET"): # Pagos de Impuestos
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS AL SENIAT'
            elif row['concepto'].startswith('PAGO BANAVIH'): # Pagos a BANAVIH
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS BANAVIH'
            elif row['concepto'].startswith('DEVUELTA PAGO PROVEEDORES'): # Reintegros
                movimientos.at[index, 'tipoOperacion'] = 'REINTEGROS TRANSFERENCIAS FALLIDAS'
            elif row['concepto'].startswith('TRANSF PROPIA BDV G200076496'): #Transferencias internas emitida a BDV
                movimientos.at[index, 'tipoOperacion'] = 'TRASFERENCIA INTERNA LTZ EMITIDA BDV'
            elif row['concepto'].startswith('PAGO RECIBIDO'): # Pagos de otros bancos
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS RECIBIDOS OTROS BANCOS'
            elif row['concepto'].startswith('TRANSF RECIBIDA BDV'): #Pagos de BDV
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS RECIBIDOS BDV'
            elif row['concepto'].startswith('PAGO PROVEEDOR CTAS PROPIAS'): #Transferencias emitidas internas bdv
                movimientos.at[index, 'tipoOperacion'] = 'TRANSFERENCIAS EMITIDA LTZ BDV'
            elif row['concepto'].startswith('COM PAGO A PROVEED CTAS PROPIA'): #Comision Transferencias emitidas internas bdv
                movimientos.at[index, 'tipoOperacion'] = 'COMISIONES TRANSFERENCIA EMITIDA LTZ BDV'
            elif row['concepto'].startswith('PAGO IVSS'): #PAgo seguro social
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS SEGURO SOCIAL'
            elif row['concepto'].startswith('DOMICILIACION'): #Pagos Domiciliados
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS DOMICILIADOS'
            elif row['concepto'].startswith('ABONO INTERESES LIQUIDACION'): #Abono intereses de liquidacion
                movimientos.at[index, 'tipoOperacion'] = 'ABONOS DE INTERESES'
            elif row['concepto'].startswith('PAGOMOVIL'): #Pago Movil
                movimientos.at[index, 'tipoOperacion'] = 'PAGOS MOVIL'
            elif row['concepto'].startswith('COBRO COMISION PAG MOVIL'): #Pago Movil
                movimientos.at[index, 'tipoOperacion'] = 'COMISIONES VARIAS'
            elif row['tipoMovimiento']=='Deposito':
                movimientos.at[index, 'tipoOperacion'] = 'DEPOSITOS RECIBIDOS'
            

        #reordenar las columnas
        movimientos = movimientos[['fecha', 'referencia', 'lote', 'concepto', 'monto', 'saldo', 'tipoOperacion']]


    return movimientos

######################
# FUNCION PARA OBTENER UN RESUMEN DE UNA CUENTA
######################
def get_resumen_cuenta(cuenta, meses_interes=None):
    movimientos = get_movimientos(cuenta,meses_interes)
    resumen_cuenta = movimientos.groupby(['tipoOperacion'])['monto'].sum()
    return resumen_cuenta


######################
# FUNCION EXPORTAR A EXCEL LA CONCILIACION DE LA CUENTA
######################
def exportar_conciliacion(cuenta, meses_interes, destino, consolidado_pagos, movimientos_huerfanos, pagos_huerfanos, transferencias_huerfanas):
    from openpyxl.styles import Font, Alignment
    from os import path
    ruta_archivo=""
    
    periodo=f"{periodo_meses(meses_interes)}"
    periodos_interesantes=['Primer Trimestre', 'Segundo Trimestre', 'Tercer Trimestre', 'Cuarto Trimestre', 'Año Completo']

    totales_consolidado = {
        'fecha': 'TOTALES',
        'ordenPago': '',
        'beneficiario': '',
        'montoOP': consolidado_pagos['montoOP'].sum(),
        'pagado': consolidado_pagos['pagado'].sum(),
        'lote': '',
        'comision': consolidado_pagos['comision'].sum(),
        'comisionRef': '',
        'retCalculada': consolidado_pagos['retCalculada'].sum(),
        'retPagada': consolidado_pagos['retPagada'].sum(),
        'retRef': '',
        'retComision': consolidado_pagos['retComision'].sum(),

    }

    consolidado_pagos = pd.concat([consolidado_pagos, pd.DataFrame(totales_consolidado, index=[0])])

    totales_movimientos_huerfanos = {
        'fecha': 'TOTALES',
        'concepto': '',
        'referencia': '',
        'lote': '',
        'monto': movimientos_huerfanos['monto'].sum(),
        'tipoOperacion':'',
    }

    movimientos_huerfanos = pd.concat([movimientos_huerfanos, pd.DataFrame(totales_movimientos_huerfanos, index=[0])])

    totales_pagos_huerfanos = {
        'fecha': 'TOTALES',
        'ordenPago': '',
        'referencia': '',
        'beneficiario': '',
        'MontoPagado': pagos_huerfanos['MontoPagado'].sum(),
        'totalRetenido': pagos_huerfanos['totalRetenido'].sum()


    }

    pagos_huerfanos = pd.concat([pagos_huerfanos, pd.DataFrame(totales_pagos_huerfanos, index=[0])])


    totales_transferencias_huerfanas = {
        'fecha': 'TOTALES',
        'CodTransf': '',
        'CuentaDestino': '',
        'Descripcion': '',
        'Referencia': '',
        'Monto': transferencias_huerfanas['Monto'].sum(),
        'OP_Relacionada': '',
        'RazonTransferencia': ''

    }
    transferencias_huerfanas = pd.concat([transferencias_huerfanas, pd.DataFrame(totales_transferencias_huerfanas, index=[0])])

    if len(meses_interes) == 1:
        ruta_archivo = path.join(destino, f'conciliacion{cuenta}_{meses_interes[0]}_{periodo}.xlsx')
        
    elif len(meses_interes) > 1:
        if periodo in periodos_interesantes:
            ruta_archivo = path.join(destino, f'conciliacion_{cuenta}_{periodo}.xlsx')
        else:
            period=''
            for mes in meses_interes:
                period=period + '_' + str(mes)
                
            ruta_archivo = path.join(destino, f'conciliacion_{cuenta}_{period[1:]}.xlsx')
    
    writer = pd.ExcelWriter(ruta_archivo, engine='openpyxl')
    consolidado_pagos.to_excel(writer, sheet_name='consolidadoPagos', index=False)
    movimientos_huerfanos.to_excel(writer, sheet_name='movimientosHuerfanos', index=False)
    pagos_huerfanos.to_excel(writer, sheet_name='pagosHuerfanos', index=False)
    transferencias_huerfanas.to_excel(writer, sheet_name='transferenciasHuerfanas', index=False)
    writer.close()

    #################################################
    ##### Dar Formato al Informe de Pagos Consolidado
    wb = openpyxl.load_workbook(ruta_archivo)
    
    
    hoja_consolidado = wb['consolidadoPagos']

    #Definir propiedades de la hoja
    hoja_consolidado.page_setup.paperSize = hoja_consolidado.PAPERSIZE_LETTER
    hoja_consolidado.page_setup.orientation = hoja_consolidado.ORIENTATION_LANDSCAPE
    hoja_consolidado.page_setup.fitToHeight = 0
    hoja_consolidado.page_setup.fitToWidth = 0
    hoja_consolidado.oddFooter.center.text = 'Pág. &P de &N'

    #Margenes
    conversor=0.3937007874 # 1cm = 0.3937007874 in
    hoja_consolidado.page_margins.top = (1.9 * conversor)
    hoja_consolidado.page_margins.bottom = (1.9 * conversor)
    hoja_consolidado.page_margins.left = (1 * conversor)
    hoja_consolidado.page_margins.right = (1 * conversor)
    hoja_consolidado.page_margins.header = (0.8 * conversor)
    hoja_consolidado.page_margins.footer = (0.8 * conversor)

    #Repetir primeras 5 filas
    hoja_consolidado.print_title_rows = '1:4'

    #Definir Ancho de Columnas
    hoja_consolidado.column_dimensions['A'].width = 9
    hoja_consolidado.column_dimensions['B'].width = 8.43
    hoja_consolidado.column_dimensions['C'].width = 32.14
    hoja_consolidado.column_dimensions['D'].width = 11
    hoja_consolidado.column_dimensions['E'].width = 11
    hoja_consolidado.column_dimensions['F'].width = 8.43
    hoja_consolidado.column_dimensions['G'].width = 9.25
    hoja_consolidado.column_dimensions['H'].width = 8.43
    hoja_consolidado.column_dimensions['I'].width = 11
    hoja_consolidado.column_dimensions['J'].width = 11
    hoja_consolidado.column_dimensions['K'].width = 8.43
    hoja_consolidado.column_dimensions['L'].width = 8.57

    #Insertar filas vacias al principios del documento
    hoja_consolidado.insert_rows(1, 3)

    #Definir altura de las filas
    for idx, row in enumerate(hoja_consolidado.iter_rows(min_row=1, max_row=hoja_consolidado.max_row, values_only=False), start=1):
        if any(cell.value is not None for cell in row):
            hoja_consolidado.row_dimensions[idx].height = 30

    #Escribir Titulo
    hoja_consolidado.merge_cells('A1:L1')
    hoja_consolidado['A1'].alignment = Alignment(horizontal='center')
    hoja_consolidado['A1'] = f'CONCILIACIÓN DE CUENTA {cuenta}'
    hoja_consolidado['A1'].font = Font(name='Calibri', size=16, bold=True)

    #Escribir Periodo
    #hoja_consolidado.row_dimensions[2].height = 5
    hoja_consolidado.merge_cells('A2:L2')
    hoja_consolidado['A2'].alignment = Alignment(horizontal='center')
    hoja_consolidado['A2'].font = Font(name='Calibri', size=14, bold=False)
    hoja_consolidado['A2'] = f'PERIODO: {periodo} 2024'

    #Definir que la fila de titulos sea negrita y que el texto se ajuste 
    for row in hoja_consolidado['A4:L4']:
        for cell in row:
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    #Definir los formatos de numero y la alineacion para cada columna de cada fila
    for row in hoja_consolidado.iter_rows(min_row=4, max_row=hoja_consolidado.max_row-1, values_only=False):
        row[0].number_format = 'dd/mm/yy'
        
        row[2].alignment = Alignment(wrap_text=True)
        row[3].number_format = '#,##0.00'
        row[4].number_format = '#,##0.00'
        
        row[6].number_format = '#,##0.00'
        
        row[8].number_format = '#,##0.00'
        row[9].number_format = '#,##0.00'
        row[10].number_format = '#,##0.00'
        row[11].number_format = '#,##0.00'

        for cell in row:
            cell.border = openpyxl.styles.borders.Border(
                    top=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    bottom=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    left=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    right=openpyxl.styles.borders.Side(style='thin', color='333333'))
        cell.font = Font(name='Calibri', size=11)
        
    # FIla de totales
    hoja_consolidado.merge_cells('A'+str(hoja_consolidado.max_row)+':C'+str(hoja_consolidado.max_row))
    #hoja_consolidado[2].alignment = Alignment(wrap_text=True)
    hoja_consolidado.cell(row=hoja_consolidado.max_row, column=4).number_format = '#,##0.00'
    hoja_consolidado.cell(row=hoja_consolidado.max_row, column=5).number_format = '#,##0.00'
    hoja_consolidado.cell(row=hoja_consolidado.max_row, column=7).number_format = '#,##0.00'
    hoja_consolidado.cell(row=hoja_consolidado.max_row, column=9).number_format = '#,##0.00'
    hoja_consolidado.cell(row=hoja_consolidado.max_row, column=10).number_format = '#,##0.00'
    hoja_consolidado.cell(row=hoja_consolidado.max_row, column=11).number_format = '#,##0.00'
    hoja_consolidado.cell(row=hoja_consolidado.max_row, column=12).number_format = '#,##0.00'
    
    wb.save(ruta_archivo)
    wb.close()
    
    #################################################
    ##### Dar Formato al Informe de Pagos Consolidado    
    wb=openpyxl.load_workbook(ruta_archivo)
    hoja_movimientos = wb['movimientosHuerfanos']

    #Definir propiedades de la hoja
    hoja_movimientos.page_setup.paperSize = hoja_movimientos.PAPERSIZE_LETTER
    hoja_movimientos.page_setup.orientation = hoja_movimientos.ORIENTATION_LANDSCAPE
    hoja_movimientos.page_setup.fitToHeight = 0
    hoja_movimientos.page_setup.fitToWidth = 0
    hoja_movimientos.oddFooter.center.text = 'Pág. &P de &N'

    #Margenes
    conversor=0.3937007874 # 1cm = 0.3937007874 in
    hoja_movimientos.page_margins.top = (1.9 * conversor)
    hoja_movimientos.page_margins.bottom = (1.9 * conversor)
    hoja_movimientos.page_margins.left = (1 * conversor)
    hoja_movimientos.page_margins.right = (1 * conversor)
    hoja_movimientos.page_margins.header = (0.8 * conversor)
    hoja_movimientos.page_margins.footer = (0.8 * conversor)

    #Repetir primeras 5 filas
    hoja_movimientos.print_title_rows = '1:4'

    #Definir Ancho de Columnas
    hoja_movimientos.column_dimensions['A'].width = 9
    hoja_movimientos.column_dimensions['B'].width = 32
    hoja_movimientos.column_dimensions['C'].width = 11
    hoja_movimientos.column_dimensions['D'].width = 11
    hoja_movimientos.column_dimensions['E'].width = 12
    hoja_movimientos.column_dimensions['F'].width = 30


    #Insertar filas vacias al principios del documento
    hoja_movimientos.insert_rows(1, 3)

    #Definir altura de las filas
    for idx, row in enumerate(hoja_movimientos.iter_rows(min_row=1, max_row=hoja_movimientos.max_row, values_only=False), start=1):
        if any(cell.value is not None for cell in row):
            hoja_movimientos.row_dimensions[idx].height = 30

    #Escribir Titulo
    hoja_movimientos.merge_cells('A1:F1')
    hoja_movimientos['A1'].alignment = Alignment(horizontal='center')
    hoja_movimientos['A1'] = f'MOVIMIENTOS HUERFANOS {cuenta}'
    hoja_movimientos['A1'].font = Font(name='Calibri', size=16, bold=True)

    #Escribir Periodo
    #hoja_movimientos.row_dimensions[2].height = 5
    hoja_movimientos.merge_cells('A2:F2')
    hoja_movimientos['A2'].alignment = Alignment(horizontal='center')
    hoja_movimientos['A2'].font = Font(name='Calibri', size=14, bold=False)
    hoja_movimientos['A2'] = f'PERIODO: {periodo} 2024'

    #Definir que la fila de titulos sea negrita y que el texto se ajuste 
    for row in hoja_movimientos['A4:F4']:
        for cell in row:
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    #Definir los formatos de numero y la alineacion para cada columna de cada fila
    for row in hoja_movimientos.iter_rows(min_row=4, max_row=hoja_movimientos.max_row-1, values_only=False):
        row[0].number_format = 'dd/mm/yy'
        row[1].alignment = Alignment(wrap_text=True)
        
        row[4].number_format = '#,##0.00'
        row[5].alignment = Alignment(wrap_text=True)
    
        for cell in row:
            cell.border = openpyxl.styles.borders.Border(
                    top=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    bottom=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    left=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    right=openpyxl.styles.borders.Side(style='thin', color='333333'))
        cell.font = Font(name='Calibri', size=11)
        
    # FIla de totales
    hoja_movimientos.merge_cells('A'+str(hoja_movimientos.max_row)+':C'+str(hoja_movimientos.max_row))
    #hoja_movimientos[2].alignment = Alignment(wrap_text=True)
    
    hoja_movimientos.cell(row=hoja_movimientos.max_row, column=5).number_format = '#,##0.00'

    wb.save(ruta_archivo)
    wb.close()
    

    ############################
    # Dar formato a la hoja de pagos huerfanos
    wb = openpyxl.load_workbook(ruta_archivo)
    hoja_pagos = wb['pagosHuerfanos']

    #Definir propiedades de la hoja
    hoja_pagos.page_setup.paperSize = hoja_pagos.PAPERSIZE_LETTER
    hoja_pagos.page_setup.orientation = hoja_pagos.ORIENTATION_LANDSCAPE
    hoja_pagos.page_setup.fitToHeight = 0
    hoja_pagos.page_setup.fitToWidth = 0
    hoja_pagos.oddFooter.center.text = 'Pág. &P de &N'

    #Margenes
    conversor=0.3937007874 # 1cm = 0.3937007874 in
    hoja_pagos.page_margins.top = (1.9 * conversor)
    hoja_pagos.page_margins.bottom = (1.9 * conversor)
    hoja_pagos.page_margins.left = (1 * conversor)
    hoja_pagos.page_margins.right = (1 * conversor)
    hoja_pagos.page_margins.header = (0.8 * conversor)
    hoja_pagos.page_margins.footer = (0.8 * conversor)

    #Repetir primeras 5 filas
    hoja_pagos.print_title_rows = '1:4'

    #Definir Ancho de Columnas
    hoja_pagos.column_dimensions['A'].width = 9
    hoja_pagos.column_dimensions['B'].width = 9
    hoja_pagos.column_dimensions['C'].width = 9
    hoja_pagos.column_dimensions['D'].width = 33
    hoja_pagos.column_dimensions['E'].width = 11
    hoja_pagos.column_dimensions['F'].width = 11

    #Insertar filas vacias al principios del documento
    hoja_pagos.insert_rows(1, 3)

    #Definir altura de las filas
    for idx, row in enumerate(hoja_pagos.iter_rows(min_row=1, max_row=hoja_pagos.max_row, values_only=False), start=1):
        if any(cell.value is not None for cell in row):
            hoja_pagos.row_dimensions[idx].height = 30

    #Escribir Titulo
    hoja_pagos.merge_cells('A1:F1')
    hoja_pagos['A1'].alignment = Alignment(horizontal='center')
    hoja_pagos['A1'] = f'Pagos Registrados Huerfanos de la cuenta {cuenta}'
    hoja_pagos['A1'].font = Font(name='Calibri', size=16, bold=True)

    #Escribir Periodo
    #hoja_pagos.row_dimensions[2].height = 5
    hoja_pagos.merge_cells('A2:F2')
    hoja_pagos['A2'].alignment = Alignment(horizontal='center')
    hoja_pagos['A2'].font = Font(name='Calibri', size=14, bold=False)
    hoja_pagos['A2'] = f'PERIODO: {periodo} 2024'

    #Definir que la fila de titulos sea negrita y que el texto se ajuste 
    for row in hoja_pagos['A4:F4']:
        for cell in row:
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    #Definir los formatos de numero y la alineacion para cada columna de cada fila
    for row in hoja_pagos.iter_rows(min_row=4, max_row=hoja_pagos.max_row-1, values_only=False):
        row[0].number_format = 'dd/mm/yy'
        
        row[3].alignment = Alignment(wrap_text=True)
        
        row[4].number_format = '#,##0.00'        
        row[5].number_format = '#,##0.00'

        for cell in row:
            cell.border = openpyxl.styles.borders.Border(
                    top=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    bottom=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    left=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    right=openpyxl.styles.borders.Side(style='thin', color='333333'))
        cell.font = Font(name='Calibri', size=11)
        
    # FIla de totales
    #hoja_pagos.merge_cells('A'+str(hoja_pagos.max_row)+':D'+str(hoja_pagos.max_row))
    #hoja_pagos[2].alignment = Alignment(wrap_text=True)
    hoja_pagos.cell(row=hoja_pagos.max_row, column=5).number_format = '#,##0.00'
    hoja_pagos.cell(row=hoja_pagos.max_row, column=6).number_format = '#,##0.00'

    wb.save(ruta_archivo)
    wb.close()


    ##################
    ## Dar formato a la hoja de transferencias huerfanas

    wb = openpyxl.load_workbook(ruta_archivo)
    hoja_transferencia = wb['transferenciasHuerfanas']

    #Definir propiedades de la hoja
    hoja_transferencia.page_setup.paperSize = hoja_transferencia.PAPERSIZE_LETTER
    hoja_transferencia.page_setup.orientation = hoja_transferencia.ORIENTATION_LANDSCAPE
    hoja_transferencia.page_setup.fitToHeight = 0
    hoja_transferencia.page_setup.fitToWidth = 0
    hoja_transferencia.oddFooter.center.text = 'Pág. &P de &N'

    #Margenes
    conversor=0.3937007874 # 1cm = 0.3937007874 in
    hoja_transferencia.page_margins.top = (1.9 * conversor)
    hoja_transferencia.page_margins.bottom = (1.9 * conversor)
    hoja_transferencia.page_margins.left = (1 * conversor)
    hoja_transferencia.page_margins.right = (1 * conversor)
    hoja_transferencia.page_margins.header = (0.8 * conversor)
    hoja_transferencia.page_margins.footer = (0.8 * conversor)

    #Repetir primeras 5 filas
    hoja_transferencia.print_title_rows = '1:4'

    #Definir Ancho de Columnas
    hoja_transferencia.column_dimensions['A'].width = 9
    hoja_transferencia.column_dimensions['B'].width = 9
    hoja_transferencia.column_dimensions['C'].width = 9
    hoja_transferencia.column_dimensions['D'].width = 36
    hoja_transferencia.column_dimensions['E'].width = 9
    hoja_transferencia.column_dimensions['F'].width = 11
    hoja_transferencia.column_dimensions['G'].width = 18
    hoja_transferencia.column_dimensions['H'].width = 18


    #Insertar filas vacias al principios del documento
    hoja_transferencia.insert_rows(1, 3)

    #Definir altura de las filas
    for idx, row in enumerate(hoja_transferencia.iter_rows(min_row=1, max_row=hoja_transferencia.max_row, values_only=False), start=1):
        if any(cell.value is not None for cell in row):
            hoja_transferencia.row_dimensions[idx].height = 30

    #Escribir Titulo
    hoja_transferencia.merge_cells('A1:H1')
    hoja_transferencia['A1'].alignment = Alignment(horizontal='center')
    hoja_transferencia['A1'] = f'TRANSFERENCIAS DESDE LA DE CUENTA {cuenta}'
    hoja_transferencia['A1'].font = Font(name='Calibri', size=16, bold=True)

    #Escribir Periodo
    #hoja_transferencia.row_dimensions[2].height = 5
    hoja_transferencia.merge_cells('A2:H2')
    hoja_transferencia['A2'].alignment = Alignment(horizontal='center')
    hoja_transferencia['A2'].font = Font(name='Calibri', size=14, bold=False)
    hoja_transferencia['A2'] = f'PERIODO: {periodo} 2024'

    #Definir que la fila de titulos sea negrita y que el texto se ajuste 
    for row in hoja_transferencia['A4:h4']:
        for cell in row:
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    #Definir los formatos de numero y la alineacion para cada columna de cada fila
    for row in hoja_transferencia.iter_rows(min_row=4, max_row=hoja_transferencia.max_row-1, values_only=False):
        row[0].number_format = 'dd/mm/yy'
        
        row[3].alignment = Alignment(wrap_text=True)
        row[5].number_format = '#,##0.00'
        row[7].alignment = Alignment(wrap_text=True)
        
        for cell in row:
            cell.border = openpyxl.styles.borders.Border(
                    top=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    bottom=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    left=openpyxl.styles.borders.Side(style='thin', color='333333'),
                    right=openpyxl.styles.borders.Side(style='thin', color='333333'))
        cell.font = Font(name='Calibri', size=11)
        
    # FIla de totales
    #hoja_transferencia.merge_cells('A'+str(hoja_transferencia.max_row)+':F'+str(hoja_transferencia.max_row))
    #hoja_transferencia[2].alignment = Alignment(wrap_text=True)
    hoja_transferencia.cell(row=hoja_transferencia.max_row, column=6).number_format = '#,##0.00'
    

    
    


    wb.save(ruta_archivo)
    wb.close()


    """
    print(f'exportar conciliacion de cuenta {cuenta}, para los meses {meses_interes} en la ubicacion {destino}')
    print("\033[92mConsolidado de pagos: \033[0m")
    print(consolidado_pagos)
    print("\033[91mMovimientos huerfanos: \033[0m")
    print(movimientos_huerfanos)
    print("\033[91mPagos huerfanos: \033[0m")
    print(pagos_huerfanos)
    print("\033[91mTransferencias huerfanas: \033[0m")
    print(transferencias_huerfanas)
    """
def periodo_meses(meses_interes):
    meses={
    'enero': 1,
    'febrero': 2,
    'marzo': 3,
    'abril': 4,
    'mayo': 5,
    'junio': 6,
    'julio': 7,
    'agosto': 8,
    'septiembre': 9,
    'octubre': 10,
    'noviembre': 11,
    'diciembre': 12
}

    if meses_interes == [1, 2, 3]:
        periodo = 'Primer Trimestre'
    elif meses_interes == [4, 5, 6]:
        periodo = 'Segundo Trimestre'
    elif meses_interes == [7, 8, 9]:
        periodo = 'Tercer Trimestre'
    elif meses_interes == [10, 11, 12]:
        periodo = 'Cuarto Trimestre'
    elif meses_interes == [1, 2,3,4,5,6,7,8,9,10,11,12]:
        periodo = 'Año Completo' 
    else:
        if len(meses_interes) == 1:
            periodo=list(meses.keys())[list(meses.values()).index(meses_interes[0])]
        else:
            periodo=''
            for mes, num in meses.items():
                if num in meses_interes:
                    periodo=periodo + mes + ', '
            periodo=periodo[:-2]

    return periodo
