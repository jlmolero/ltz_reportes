import pandas as pd
import datetime as dt
import openpyxl
import yaml
import re
from os import path

with open('./config.yaml', 'r') as f:
    config = yaml.safe_load(f)



#ruta_pagos='./data/RegistroPagos.xlsx'
ruta_pagos = config['origen_pagos']
#ruta_pagos = r"\\dc2\administracion\2024\CUADROS\Registro de Pagos.xlsx"

# Define un diccionario que mapea los valores de la columna Cuenta a los valores de la columna FuenteDeFinanciamiento
cuenta_fuente_financiamiento = {
    '4363': 'Recursos por Operaciones',
    '6597': 'Recursos por Operaciones',
    '0171': 'Situado Constitucional',
    'PATRIA': 'Situado Constitucional'
}

partida_descripcion = {
    '4.01':'GASTOS DE PERSONAL',
    '4.02':'MATERIALES, SUMINISTROS Y MERCANCÍAS',
    '4.03':'SERVICIOS NO PERSONALES',
    '4.04':'ACTIVOS REALES',
    '4.05':'ACTIVOS  FINANCIEROS',
    '4.06':'GASTOS DE DEFENSA Y SEGURIDAD DEL ESTADO',
    '4.07':'TRANSFERENCIAS Y DONACIONES',
    '4.08':'OTROS GASTOS',
    '4.09':'ASIGNACIONES NO DISTRIBUIDAS',
    '4.10':'SERVICIO DE LA DEUDA PÚBLICA',
    '4.11':'DISMINUCIÓN DE PASIVOS',
    '4.12':'DISMINUCIÓN DEL PATRIMONIO',
    '4.98':'RECTIFICACIONES AL PRESUPUESTO'
}

def periodo_meses(meses_interes):
    meses={
    'enero': 1,
    'febrero': 2,
    'marzo': 3,
    'abril': 4,
    'mayo': 5,
    'junio': 6,
    'julio': 7,
    'agosto': 8,
    'septiembre': 9,
    'octubre': 10,
    'noviembre': 11,
    'diciembre': 12
}

    if meses_interes == [1, 2, 3]:
        periodo = 'Primer Trimestre'
    elif meses_interes == [4, 5, 6]:
        periodo = 'Segundo Trimestre'
    elif meses_interes == [7, 8, 9]:
        periodo = 'Tercer Trimestre'
    elif meses_interes == [10, 11, 12]:
        periodo = 'Cuarto Trimestre'
    elif meses_interes == [1, 2,3,4,5,6,7,8,9,10,11,12]:
        periodo = 'Año Completo' 
    else:
        if len(meses_interes) == 1:
            periodo=list(meses.keys())[list(meses.values()).index(meses_interes[0])]
        else:
            periodo=''
            for mes, num in meses.items():
                if num in meses_interes:
                    periodo=periodo + mes + ', '
            periodo=periodo[:-2]

    return periodo

def gastos_periodo(meses_interes):
    # ejecucion_gasto.py

    """
    meses={
        'enero': 1,
        'febrero': 2,
        'marzo': 3,
        'abril': 4,
        'mayo': 5,
        'junio': 6,
        'julio': 7,
        'agosto': 8,
        'septiembre': 9,
        'octubre': 10,
        'noviembre': 11,
        'diciembre': 12
    }

    if meses_interes == [1, 2, 3]:
        periodo = 'Primer Trimestre'
    elif meses_interes == [4, 5, 6]:
        periodo = 'Segundo Trimestre'
    elif meses_interes == [7, 8, 9]:
        periodo = 'Tercer Trimestre'
    elif meses_interes == [10, 11, 12]:
        periodo = 'Cuarto Trimestre'
    else:
        periodo = meses_interes
    """
    datos = pd.read_excel(ruta_pagos, sheet_name="Pagos") # Importar datos
    
    #Eliminar registros
    datos= datos[datos['NroFactura']!='ANULADA'] # Eliminar ordenes de pago anuladas
    datos=datos[datos['OrdenPago'].notnull()] # Eliminar ordenes de pago vacías
    datos = datos[datos['CodigoPartida']!='4.03.18.01.00'] # Eliminar registros de ordemes de pagos de retencion de IVA
    datos = datos[datos['CodigoPartida']!='4.03.18.99.00'] # Eliminar registros de ordemes de pagos de retencion de islr y sedatez


    # Seleccionar solo los registros que corresponden al periodo deseado
    datos = datos[datos['Fecha'].dt.month.isin(meses_interes)]

    #Sustituir valores - en la columna IVA
    #datos['IVA'] = datos['IVA'].replace('-', 0)
    datos['IVA'] = datos['IVA'].apply(lambda x: 0 if x == '-' else x)


    #crear una columna que contendra las diferentes fuentes de financimiento segun la cuenta de donde sale los recursos
    fuente_financiamiento = datos['Cuenta'].map(cuenta_fuente_financiamiento)
    #Asignar la fuente de financiamiento a las columnas donde ya venga especificada
    datos = datos.assign(FuenteDeFinanciamiento=datos['FuenteDeFinanciamiento'].fillna(fuente_financiamiento))
    

    #Seleccionar las columnas deseadas para la tabla de datos
    datos=datos[['Fecha', 'Cuenta', 'OrdenPago', 'Beneficiario', 'CodigoPartida', 'DescripcionPartida', 'Descripcion', 'MontoSinIVA', 'IVA', 'FuenteDeFinanciamiento']]
    
    
    return datos

def informe_ejecucion_gasto(datos, meses_interes):
    iva_periodo= datos[datos['Fecha'].dt.month.isin(meses_interes)].pivot_table(values='IVA', index=None, columns='FuenteDeFinanciamiento', aggfunc='sum').infer_objects(copy=False).fillna(0).reset_index()
    #Reindexamos las columnas para garantizar que tenga valores en todas las fuentes de financiamiento
    iva_periodo = iva_periodo.reindex(columns=['Recursos por Operaciones', 'Situado Constitucional'], fill_value=0)

    fila_iva_mes = pd.DataFrame({'CodigoPartida':['4.03.18.01.00'], 'DescripcionPartida':['Impuesto al valor Agregado'], 'Recursos por Operaciones':[iva_periodo.loc[0, 'Recursos por Operaciones']], 'Situado Constitucional':[iva_periodo.loc[0, 'Situado Constitucional']]}).fillna(0)

    informe_gasto_ejecutado = datos.groupby(['CodigoPartida', 'DescripcionPartida', 'FuenteDeFinanciamiento'])['MontoSinIVA'].sum().unstack('FuenteDeFinanciamiento').fillna(0).rename_axis(None, axis=1).reset_index()
    # Reindexamos las columnas para garantizar que tenga valores en todas las fuentes de financiamiento
    informe_gasto_ejecutado = informe_gasto_ejecutado.reindex(columns=['CodigoPartida', 'DescripcionPartida', 'Recursos por Operaciones', 'Situado Constitucional'], fill_value=0)

    informe_gasto_ejecutado = pd.concat([informe_gasto_ejecutado, fila_iva_mes], ignore_index=True).sort_values(by='CodigoPartida')

    # Ahora voy a agregar una columna con el valor total de 'MontoSinIVA' para cada agrupación
    informe_gasto_ejecutado['Total'] = informe_gasto_ejecutado.iloc[:, 2:].sum(axis=1)

    #Calcular subtotales    
    informe_gasto_ejecutado_subtotales = informe_gasto_ejecutado.groupby(informe_gasto_ejecutado['CodigoPartida'].str[:4]).agg({'Recursos por Operaciones':'sum', 'Situado Constitucional':'sum', 'Total': 'sum'}).rename_axis('CodigoPartida').reset_index()
    informe_gasto_ejecutado_subtotales['CodigoPartida'] = informe_gasto_ejecutado_subtotales['CodigoPartida']# + '.00.00.00'
    informe_gasto_ejecutado_subtotales['DescripcionPartida'] = informe_gasto_ejecutado_subtotales['CodigoPartida'].map(partida_descripcion)

    #fila de totales generales
    fila_totales_generales = pd.DataFrame({'CodigoPartida':['TOTALES'], 'DescripcionPartida':[''], 
                                      'Recursos por Operaciones':[informe_gasto_ejecutado['Recursos por Operaciones'].sum()], 
                                      'Situado Constitucional':[informe_gasto_ejecutado['Situado Constitucional'].sum()],
                                      'Total':[informe_gasto_ejecutado['Total'].sum()]})

    # Agregar subtotales y totales generales
    informe_gasto_ejecutado = pd.concat([informe_gasto_ejecutado, informe_gasto_ejecutado_subtotales], ignore_index=True).sort_values(by='CodigoPartida')
    informe_gasto_ejecutado = pd.concat([informe_gasto_ejecutado, fila_totales_generales], ignore_index=True).sort_values(by='CodigoPartida')

 
    #return informe_gasto_ejecutado
    return informe_gasto_ejecutado

def informe_ejecucion_cuenta(datos, meses_interes):
    
    iva_periodo= datos[datos['Fecha'].dt.month.isin(meses_interes)].pivot_table(values='IVA', index=None, columns='Cuenta', aggfunc='sum').infer_objects(copy=False).fillna(0).reset_index()
    # Reindexamos el dataframe para garantizar que tenga valores para todas las cuentas
    iva_periodo = iva_periodo.reindex(columns=['0171', '2633', '4363', '6597', 'PATRIA'], fill_value=0.0)

    
    fila_iva_mes = pd.DataFrame({'CodigoPartida':['4.03.18.01.00'], 'DescripcionPartida':['Impuesto al valor Agregado'],
                                 '0171':[iva_periodo.loc[0, '0171']],
                                 '2633':[iva_periodo.loc[0, '2633']],
                                 '4363':[iva_periodo.loc[0,'4363']],
                                 '6597':[iva_periodo.loc[0,'6597']],
                                 'PATRIA':[iva_periodo.loc[0,'PATRIA']]
                                 }).fillna(0)

    
    informe_gasto_ejecutado = datos.groupby(['CodigoPartida', 'DescripcionPartida', 'Cuenta'])['MontoSinIVA'].sum().unstack('Cuenta').fillna(0).rename_axis(None, axis=1).reset_index()
    # Reindexamos el dataframe para garantizar que tenga valores para todas las Cuentas
    informe_gasto_ejecutado = informe_gasto_ejecutado.reindex(columns=['CodigoPartida', 'DescripcionPartida', '0171', '2633', '4363', '6597', 'PATRIA'], fill_value=0.0)
    informe_gasto_ejecutado = pd.concat([informe_gasto_ejecutado, fila_iva_mes], ignore_index=True).sort_values(by='CodigoPartida')
    
    # Ahora voy a agregar una columna con el valor total de 'MontoSinIVA' para cada agrupación
    informe_gasto_ejecutado['Total'] = informe_gasto_ejecutado.iloc[:, 2:].sum(axis=1)

    #Calcular subtotales    
    informe_gasto_ejecutado_subtotales = informe_gasto_ejecutado.groupby(informe_gasto_ejecutado['CodigoPartida'].str[:4]).agg({'0171':'sum', '2633':'sum', '4363':'sum', '6597':'sum', 'PATRIA':'sum', 'Total': 'sum'}).rename_axis('CodigoPartida').reset_index()
    
    #informe_gasto_ejecutado_subtotales['CodigoPartida'] = informe_gasto_ejecutado_subtotales['CodigoPartida']# + '.00.00.00'
    informe_gasto_ejecutado_subtotales['DescripcionPartida'] = informe_gasto_ejecutado_subtotales['CodigoPartida'].map(partida_descripcion)
    informe_gasto_ejecutado = pd.concat([informe_gasto_ejecutado, informe_gasto_ejecutado_subtotales], ignore_index=True).sort_values(by='CodigoPartida')

    
    return informe_gasto_ejecutado

def ejecucion_gasto_excel(informe_gasto_ejecutado, meses_interes, ruta):
    from openpyxl.styles import Font, Alignment
    from os import path
    ruta_archivo=""
    
    periodo=f"{periodo_meses(meses_interes)}"
    periodos_interesantes=['Primer Trimestre', 'Segundo Trimestre', 'Tercer Trimestre', 'Cuarto Trimestre', 'Año Completo']

    ##########################
    ######### Exportar a Excel
    ##########################
    
    if len(meses_interes) == 1:
        ruta_archivo = path.join(ruta, f'informe_gasto_ejecutado_{meses_interes[0]}_{periodo}.xlsx')
        
    elif len(meses_interes) > 1:
        if periodo in periodos_interesantes:
            ruta_archivo = path.join(ruta, f'informe_gasto_ejecutado_{periodo}.xlsx')
        else:
            period=''
            for mes in meses_interes:
                period=period + '_' + str(mes)
                
            ruta_archivo = path.join(ruta, f'informe_gasto_ejecutado_{period[1:]}.xlsx')
    
    informe_gasto_ejecutado.to_excel(ruta_archivo, index=False)


    ##########################
    ####### Dar Formato al Iinforme
    ##########################
    
    wb = openpyxl.load_workbook(ruta_archivo)
    sheet = wb.active

    #Definir propiedades de la hoja
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.fitToWidth = 0
    sheet.oddFooter.center.text = 'Pág. &P de &N'
    #Margenes
    conversor=0.3937007874 # 1cm = 0.3937007874 in
    sheet.page_margins.top = (1.9 * conversor)
    sheet.page_margins.bottom = (1.9 * conversor)
    sheet.page_margins.left = (0.6 * conversor)
    sheet.page_margins.right = (0.6 * conversor)
    sheet.page_margins.header = (0.8 * conversor)
    sheet.page_margins.footer = (0.8 * conversor)
    #Repetir primeras 5 filas
    sheet.print_title_rows = '1:5'

    #Definir Ancho de Columnas
    sheet.column_dimensions['A'].width = 13.71
    sheet.column_dimensions['B'].width = 44.14
    sheet.column_dimensions['C'].width = 13.71
    sheet.column_dimensions['D'].width = 13.71
    sheet.column_dimensions['E'].width = 13.71

    #Insertar filas vacias al principios del documento
    sheet.insert_rows(1, 4)


    #Definir altura de las filas
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False), start=1):
        if any(cell.value is not None for cell in row):
            sheet.row_dimensions[idx].height = 30

    #Escribir Titulo
    sheet.merge_cells('A1:E1')
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['A1'] = 'INFORME DE GASTO EJECUTADO'
    sheet['A1'].font = Font(name='Calibri', size=16, bold=True)

    #Escribir Periodo
    sheet.row_dimensions[2].height = 5
    sheet.merge_cells('D3:E3')
    sheet['C3'].alignment = Alignment(horizontal='right')
    sheet['C3'].font = Font(name='Calibri', size=12, bold=False)
    sheet['C3'] = 'PERIODO: '
    sheet['D3'].font = Font(name='Calibri', size=12, bold=True)
    sheet['D3'] = f"{periodo} 2024"

    #Definir que la fila de titulos sea negrita y que el texto se ajuste 
    for row in sheet['A5:E5']:
        for cell in row:
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, values_only=False):
        #Filas de Subtotales de Partidas
        if isinstance(row[0].value, str) and re.match(r'^\d\.\d{2}$', row[0].value): #len(row[0].value) == 4:
            row[1].alignment = Alignment(wrap_text=True)
            row[2].number_format = '#,##0.00'
            row[3].number_format = '#,##0.00'
            row[4].number_format = '#,##0.00'
            
            for cell in row:
                cell.border = openpyxl.styles.borders.Border(
                    top=openpyxl.styles.borders.Side(style='thin'),
                    bottom=openpyxl.styles.borders.Side(style='thin'),
                    left=openpyxl.styles.borders.Side(style='thin'),
                    right=openpyxl.styles.borders.Side(style='thin'))
                cell.fill = openpyxl.styles.PatternFill(fgColor='D9D9D9', fill_type='solid')
                cell.font = Font(name='Calibri', size=12, bold=True)
        #filas Normales
        elif isinstance(row[0].value, str) and re.match(r'^\d\.\d{2}\.\d{2}\.\d{2}\.\d{2}$', row[0].value):
            row[1].alignment = Alignment(wrap_text=True)
            row[2].number_format = '#,##0.00'
            row[3].number_format = '#,##0.00'
            row[4].number_format = '#,##0.00'
            
            for cell in row:
                cell.border = openpyxl.styles.borders.Border(
                    top=openpyxl.styles.borders.Side(style='thin'),
                    bottom=openpyxl.styles.borders.Side(style='thin'),
                    left=openpyxl.styles.borders.Side(style='thin'),
                    right=openpyxl.styles.borders.Side(style='thin'))
                cell.font = Font(name='Calibri', size=11)

        #Fila Totales
        elif isinstance(row[0].value, str) and row[0].value=='TOTALES':
            row[2].number_format = '#,##0.00'
            row[3].number_format = '#,##0.00'
            row[4].number_format = '#,##0.00'
            
            for cell in row:
                cell.border = openpyxl.styles.borders.Border(
                    top=openpyxl.styles.borders.Side(style='thin'),
                    bottom=openpyxl.styles.borders.Side(style='thin'),
                    left=openpyxl.styles.borders.Side(style='thin'),
                    right=openpyxl.styles.borders.Side(style='thin'))
                cell.fill = openpyxl.styles.PatternFill(fgColor='404040', fill_type='solid')
                cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')



    wb.save(ruta_archivo)
    return True

    